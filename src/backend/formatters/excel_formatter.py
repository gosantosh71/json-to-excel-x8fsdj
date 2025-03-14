"""
Provides Excel workbook-level formatting and generation capabilities for the JSON to Excel Conversion Tool.
This module handles the creation, formatting, and saving of Excel workbooks, serving as a key component 
in the Excel generation feature of the application.
"""

import os  # v: built-in
from typing import Dict, List, Optional, Union, Any, ByteString  # v: built-in
import io  # v: built-in

import openpyxl  # v: 3.1.0+
import pandas  # v: 1.5.0+

from ..logger import get_logger
from .column_formatter import ColumnFormatter
from ..models.excel_options import ExcelOptions
from ..models.json_data import JSONData
from ..constants import EXCEL_CONSTANTS
from ..exceptions import ExcelGenerationException, ExcelLimitExceededException

# Initialize module logger
logger = get_logger(__name__)


def create_workbook(properties: Dict[str, Any] = None) -> openpyxl.Workbook:
    """
    Creates a new Excel workbook with optional properties.
    
    Args:
        properties: Dictionary of workbook properties like title, author, etc.
        
    Returns:
        A new Excel workbook instance
    """
    workbook = openpyxl.Workbook()
    
    # Apply properties if provided
    if properties:
        for key, value in properties.items():
            if hasattr(workbook.properties, key):
                setattr(workbook.properties, key, value)
    
    logger.debug("Created new Excel workbook")
    return workbook


def save_workbook(workbook: openpyxl.Workbook, file_path: str) -> bool:
    """
    Saves an Excel workbook to the specified file path.
    
    Args:
        workbook: The Excel workbook to save
        file_path: Path where the workbook will be saved
        
    Returns:
        True if the workbook was saved successfully
        
    Raises:
        ExcelGenerationException: If an error occurs while saving the workbook
    """
    try:
        # Ensure directory exists
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
            logger.debug(f"Created directory: {directory}")
        
        # Save the workbook
        workbook.save(file_path)
        logger.info(f"Excel workbook saved successfully to {file_path}")
        return True
    
    except PermissionError as e:
        error_msg = f"Permission denied while saving Excel file to {file_path}"
        logger.error(error_msg)
        raise ExcelGenerationException(
            message=error_msg, 
            context={"file_path": file_path, "error": str(e)}
        )
    
    except Exception as e:
        error_msg = f"Error saving Excel file to {file_path}: {str(e)}"
        logger.error(error_msg)
        raise ExcelGenerationException(
            message=error_msg, 
            context={"file_path": file_path, "error": str(e)}
        )


def workbook_to_bytes(workbook: openpyxl.Workbook) -> ByteString:
    """
    Converts an Excel workbook to a bytes object for in-memory operations.
    
    Args:
        workbook: The Excel workbook to convert
        
    Returns:
        The workbook content as bytes
        
    Raises:
        ExcelGenerationException: If an error occurs while converting the workbook
    """
    try:
        # Create a bytes buffer
        output = io.BytesIO()
        
        # Save the workbook to the buffer
        workbook.save(output)
        
        # Get the bytes content and reset buffer position
        output.seek(0)
        content = output.getvalue()
        
        logger.debug("Converted Excel workbook to bytes")
        return content
    
    except Exception as e:
        error_msg = f"Error converting Excel workbook to bytes: {str(e)}"
        logger.error(error_msg)
        raise ExcelGenerationException(message=error_msg)


def validate_excel_limits(df: pandas.DataFrame) -> bool:
    """
    Validates that a DataFrame doesn't exceed Excel's row and column limits.
    
    Args:
        df: The pandas DataFrame to validate
        
    Returns:
        True if the DataFrame is within Excel limits
        
    Raises:
        ExcelLimitExceededException: If the DataFrame exceeds Excel's limits
    """
    rows, cols = df.shape
    max_rows = EXCEL_CONSTANTS["MAX_ROWS"]
    max_cols = EXCEL_CONSTANTS["MAX_COLUMNS"]
    
    if rows > max_rows:
        error_msg = f"Data exceeds Excel's row limit: {rows} rows (maximum: {max_rows})"
        logger.error(error_msg)
        raise ExcelLimitExceededException(
            row_count=rows,
            max_rows=max_rows,
            context={"dataframe_shape": f"{rows}x{cols}"}
        )
    
    if cols > max_cols:
        error_msg = f"Data exceeds Excel's column limit: {cols} columns (maximum: {max_cols})"
        logger.error(error_msg)
        raise ExcelLimitExceededException(
            column_count=cols,
            max_columns=max_cols,
            context={"dataframe_shape": f"{rows}x{cols}"}
        )
    
    logger.debug(f"DataFrame dimensions ({rows} rows, {cols} columns) are within Excel limits")
    return True


def dataframe_to_excel(
    df: pandas.DataFrame, 
    workbook: openpyxl.Workbook, 
    sheet_name: str = None,
    options: ExcelOptions = None
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Converts a pandas DataFrame to an Excel worksheet in the given workbook.
    
    Args:
        df: The pandas DataFrame to convert
        workbook: The workbook to add the worksheet to
        sheet_name: The name for the worksheet (optional)
        options: Excel formatting options (optional)
        
    Returns:
        The created worksheet
        
    Raises:
        ExcelLimitExceededException: If the DataFrame exceeds Excel's limits
        ExcelGenerationException: If an error occurs during conversion
    """
    # Use default sheet name if not provided
    sheet_name = sheet_name or EXCEL_CONSTANTS["DEFAULT_SHEET_NAME"]
    
    # Validate DataFrame against Excel limits
    validate_excel_limits(df)
    
    try:
        # Check if the default sheet exists and is empty
        if sheet_name in workbook.sheetnames:
            # Use existing sheet
            worksheet = workbook[sheet_name]
        else:
            # Create a new sheet or use the active one if it's the default sheet
            if workbook.active.title == "Sheet" and not workbook.active._cells:
                # Rename the default empty sheet
                worksheet = workbook.active
                worksheet.title = sheet_name
            else:
                # Create a new sheet
                worksheet = workbook.create_sheet(title=sheet_name)
        
        # Write the header row (column names)
        for col_idx, col_name in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=str(col_name))
        
        # Write the data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):  # Start from row 2 (after header)
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply column formatting if options are provided
        if options:
            column_formatter = ColumnFormatter(options)
            column_formatter.format_worksheet(worksheet)
        
        logger.info(f"Converted DataFrame with {len(df)} rows to Excel worksheet '{sheet_name}'")
        return worksheet
        
    except (ExcelLimitExceededException, ExcelGenerationException):
        # Re-raise known exceptions
        raise
    except Exception as e:
        error_msg = f"Error converting DataFrame to Excel: {str(e)}"
        logger.error(error_msg)
        raise ExcelGenerationException(
            message=error_msg, 
            context={"sheet_name": sheet_name, "error": str(e)}
        )


class ExcelFormatter:
    """
    A class that handles Excel workbook creation, formatting, and generation from JSON data.
    """
    
    def __init__(self, options: ExcelOptions, json_data: Optional[JSONData] = None):
        """
        Initializes a new ExcelFormatter instance with the provided options and JSON data.
        
        Args:
            options: Excel formatting options
            json_data: Optional JSON data for additional context
        """
        self._options = options
        self._json_data = json_data
        self._logger = get_logger(__name__)
        self._logger.debug("ExcelFormatter initialized")
    
    def format_excel(self, workbook: openpyxl.Workbook) -> openpyxl.Workbook:
        """
        Formats an Excel workbook according to the specified options.
        
        Args:
            workbook: The workbook to format
            
        Returns:
            The formatted workbook
        """
        # Apply workbook properties if specified
        if self._options.workbook_properties:
            for key, value in self._options.workbook_properties.items():
                if hasattr(workbook.properties, key):
                    setattr(workbook.properties, key, value)
        
        # Format each worksheet in the workbook
        for worksheet in workbook.worksheets:
            # Apply column formatting 
            column_formatter = ColumnFormatter(self._options, self._json_data)
            column_formatter.format_worksheet(worksheet)
        
        self._logger.info("Applied formatting to Excel workbook")
        return workbook
    
    def generate_excel_file(self, df: pandas.DataFrame, output_path: str) -> str:
        """
        Generates an Excel file from a pandas DataFrame and saves it to the specified path.
        
        Args:
            df: The DataFrame to convert to Excel
            output_path: Path where the Excel file should be saved
            
        Returns:
            The path to the generated Excel file
            
        Raises:
            ExcelGenerationException: If an error occurs during Excel generation
            ExcelLimitExceededException: If the DataFrame exceeds Excel's limits
        """
        try:
            # Create a new workbook
            workbook_properties = self._options.workbook_properties or {}
            workbook = create_workbook(workbook_properties)
            
            # Convert the DataFrame to Excel
            sheet_name = self._options.sheet_name or EXCEL_CONSTANTS["DEFAULT_SHEET_NAME"]
            dataframe_to_excel(df, workbook, sheet_name, self._options)
            
            # Apply additional formatting
            self.format_excel(workbook)
            
            # Save the workbook
            save_workbook(workbook, output_path)
            
            self._logger.info(f"Successfully generated Excel file at {output_path}")
            return output_path
            
        except (ExcelLimitExceededException, ExcelGenerationException):
            # Re-raise known exceptions
            raise
        except Exception as e:
            error_msg = f"Error generating Excel file: {str(e)}"
            self._logger.error(error_msg)
            raise ExcelGenerationException(
                message=error_msg, 
                context={"output_path": output_path, "error": str(e)}
            )
    
    def generate_excel_bytes(self, df: pandas.DataFrame) -> ByteString:
        """
        Generates an Excel file from a pandas DataFrame and returns it as bytes.
        
        Args:
            df: The DataFrame to convert to Excel
            
        Returns:
            The Excel file as bytes
            
        Raises:
            ExcelGenerationException: If an error occurs during Excel generation
            ExcelLimitExceededException: If the DataFrame exceeds Excel's limits
        """
        try:
            # Create a new workbook
            workbook_properties = self._options.workbook_properties or {}
            workbook = create_workbook(workbook_properties)
            
            # Convert the DataFrame to Excel
            sheet_name = self._options.sheet_name or EXCEL_CONSTANTS["DEFAULT_SHEET_NAME"]
            dataframe_to_excel(df, workbook, sheet_name, self._options)
            
            # Apply additional formatting
            self.format_excel(workbook)
            
            # Convert to bytes
            excel_bytes = workbook_to_bytes(workbook)
            
            self._logger.info("Successfully generated Excel file as bytes")
            return excel_bytes
            
        except (ExcelLimitExceededException, ExcelGenerationException):
            # Re-raise known exceptions
            raise
        except Exception as e:
            error_msg = f"Error generating Excel bytes: {str(e)}"
            self._logger.error(error_msg)
            raise ExcelGenerationException(
                message=error_msg, 
                context={"error": str(e)}
            )
    
    def add_data_sheet(
        self, 
        workbook: openpyxl.Workbook, 
        df: pandas.DataFrame, 
        sheet_name: Optional[str] = None
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Adds a data sheet to an existing workbook from a pandas DataFrame.
        
        Args:
            workbook: The workbook to add the sheet to
            df: The DataFrame to convert to a worksheet
            sheet_name: Optional name for the sheet (uses options or default if not provided)
            
        Returns:
            The created worksheet
            
        Raises:
            ExcelLimitExceededException: If the DataFrame exceeds Excel's limits
            ExcelGenerationException: If an error occurs during conversion
        """
        # Determine sheet name
        sheet_name = sheet_name or self._options.sheet_name or EXCEL_CONSTANTS["DEFAULT_SHEET_NAME"]
        
        try:
            # Create a unique sheet name if one already exists with this name
            base_name = sheet_name
            counter = 1
            while sheet_name in workbook.sheetnames:
                sheet_name = f"{base_name}_{counter}"
                counter += 1
            
            # Create a new sheet
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Convert DataFrame to Excel (reuse dataframe_to_excel but with the new worksheet)
            # Write the header row (column names)
            for col_idx, col_name in enumerate(df.columns, 1):
                worksheet.cell(row=1, column=col_idx, value=str(col_name))
            
            # Write the data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):  # Start from row 2 (after header)
                for col_idx, value in enumerate(row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply formatting
            column_formatter = ColumnFormatter(self._options, self._json_data)
            column_formatter.format_worksheet(worksheet)
            
            self._logger.info(f"Added data sheet '{sheet_name}' with {len(df)} rows to workbook")
            return worksheet
            
        except (ExcelLimitExceededException, ExcelGenerationException):
            # Re-raise known exceptions
            raise
        except Exception as e:
            error_msg = f"Error adding data sheet to workbook: {str(e)}"
            self._logger.error(error_msg)
            raise ExcelGenerationException(
                message=error_msg, 
                context={"sheet_name": sheet_name, "error": str(e)}
            )