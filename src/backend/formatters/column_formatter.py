"""
Provides column-specific formatting capabilities for Excel worksheets during the JSON to Excel conversion process.
This module handles formatting of column headers, adjustment of column widths based on content, application of 
data type-specific formatting, and sanitization of cell content to prevent formula injection.
"""

import re  # v: built-in
from typing import Dict, List, Optional, Any, Union  # v: built-in

import openpyxl  # v: 3.1.0+
import pandas  # v: 1.5.0+

from ..logger import get_logger
from ..models.excel_options import ExcelOptions
from ..models.json_data import JSONData
from ..constants import EXCEL_CONSTANTS

# Initialize module globals
logger = get_logger(__name__)
HEADER_STYLE = {"font": {"bold": True}, "fill": {"patternType": "solid", "fgColor": "E0E0E0"}, "border": {"outline": True}}
FORMULA_PATTERN = re.compile(r'^[=\+\-@]')
DATA_TYPE_FORMATS = {
    "int": "0",
    "float": "0.00",
    "datetime": "yyyy-mm-dd hh:mm:ss",
    "date": "yyyy-mm-dd",
    "percentage": "0.00%",
    "currency": "$#,##0.00"
}


def format_column_headers(worksheet: openpyxl.worksheet.worksheet.Worksheet, apply_formatting: bool) -> None:
    """
    Formats the header row of an Excel worksheet with bold text, background color, and borders.
    
    Args:
        worksheet: The openpyxl worksheet to format
        apply_formatting: Boolean flag to determine if formatting should be applied
        
    Returns:
        None: Modifies worksheet in-place
    """
    if not apply_formatting:
        logger.debug("Header formatting skipped (apply_formatting=False)")
        return
    
    # Get dimensions to identify the header row
    if worksheet.max_row < 1:
        logger.warning("Worksheet is empty, no headers to format")
        return
    
    # Apply styles to the header row (first row)
    for cell in worksheet[1]:
        # Apply bold font
        cell.font = openpyxl.styles.Font(bold=HEADER_STYLE["font"]["bold"])
        
        # Apply background fill
        fill_color = HEADER_STYLE["fill"]["fgColor"]
        cell.fill = openpyxl.styles.PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type=HEADER_STYLE["fill"]["patternType"]
        )
        
        # Apply border if specified
        if HEADER_STYLE["border"]["outline"]:
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            cell.border = thin_border
            
        # Center align the header text
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    logger.debug("Applied formatting to header row")


def adjust_column_widths(worksheet: openpyxl.worksheet.worksheet.Worksheet, auto_adjust: bool) -> None:
    """
    Adjusts column widths based on content to ensure all data is visible.
    
    Args:
        worksheet: The openpyxl worksheet to adjust
        auto_adjust: Boolean flag to determine if auto-adjustment should be applied
        
    Returns:
        None: Modifies worksheet in-place
    """
    if not auto_adjust:
        logger.debug("Column width adjustment skipped (auto_adjust=False)")
        return
    
    default_width = EXCEL_CONSTANTS.get("DEFAULT_COLUMN_WIDTH", 10)
    
    # Determine column count
    max_column = worksheet.max_column
    
    for column_idx in range(1, max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(column_idx)
        max_length = 0
        
        # Check each cell in the column
        for cell in worksheet[col_letter]:
            if cell.value:
                # Calculate length based on string representation
                cell_text = str(cell.value)
                # Adjust length based on data type and expected rendering width
                if isinstance(cell.value, (int, float)):
                    # Numbers typically need less width
                    cell_length = len(cell_text) * 1.1
                elif isinstance(cell.value, str) and len(cell_text) > 20:
                    # Long strings can be wrapped
                    cell_length = min(len(cell_text), 40)
                else:
                    cell_length = len(cell_text) * 1.2
                    
                max_length = max(max_length, cell_length)
        
        # Set column width, with minimum and maximum constraints
        adjusted_width = max(default_width, min(max_length, 50))  # Cap at 50 characters
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    logger.debug(f"Adjusted widths for {max_column} columns")


def apply_data_formatting(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, 
    column_formats: Dict[str, str],
    apply_formatting: bool
) -> None:
    """
    Applies data type-specific formatting to cells based on content type.
    
    Args:
        worksheet: The openpyxl worksheet to format
        column_formats: Dictionary mapping column names to Excel format strings
        apply_formatting: Boolean flag to determine if formatting should be applied
        
    Returns:
        None: Modifies worksheet in-place
    """
    if not apply_formatting:
        logger.debug("Data formatting skipped (apply_formatting=False)")
        return
    
    if not column_formats:
        # No explicit formats provided, detect from data
        logger.debug("No column formats provided, detecting from data")
        detected_formats = detect_column_data_types(worksheet)
        column_formats = {
            col_name: DATA_TYPE_FORMATS.get(data_type, "General") 
            for col_name, data_type in detected_formats.items()
        }
    
    # Create a mapping of column indices to column names
    column_names = {}
    for column_idx in range(1, worksheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(column_idx)
        header_cell = worksheet[f"{col_letter}1"]
        if header_cell.value:
            column_names[column_idx] = str(header_cell.value)
    
    # Apply formats to data cells (skipping header row)
    for row_idx in range(2, worksheet.max_row + 1):
        for column_idx in range(1, worksheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(column_idx)
            cell = worksheet[f"{col_letter}{row_idx}"]
            
            # Get column name for this cell
            column_name = column_names.get(column_idx)
            if column_name and column_name in column_formats:
                # Apply number format from column_formats
                cell.number_format = column_formats[column_name]
                
                # Sanitize cell content to prevent formula injection
                if isinstance(cell.value, str):
                    cell.value = sanitize_cell_content(cell.value)
    
    logger.debug(f"Applied data formatting to {len(column_formats)} columns")


def sanitize_cell_content(value: Any) -> Any:
    """
    Sanitizes cell content to prevent Excel formula injection attacks.
    
    Args:
        value: The cell value to sanitize
        
    Returns:
        The sanitized cell value
    """
    # Only need to sanitize string values
    if not isinstance(value, str):
        return value
    
    # Check if the string starts with formula characters
    if FORMULA_PATTERN.match(value):
        # Prefix with a single quote to prevent formula execution
        return f"'{value}"
    
    return value


def detect_column_data_types(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, str]:
    """
    Analyzes worksheet data to detect the most appropriate data type for each column.
    
    Args:
        worksheet: The openpyxl worksheet to analyze
        
    Returns:
        A dictionary mapping column names to detected data types
    """
    # Initialize column data types
    column_types = {}
    
    # Get header row to determine column names
    column_names = {}
    for column_idx in range(1, worksheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(column_idx)
        header_cell = worksheet[f"{col_letter}1"]
        if header_cell.value:
            column_names[column_idx] = str(header_cell.value)
            column_types[str(header_cell.value)] = "string"  # Default type
    
    # Analyze data rows (skipping header)
    for row_idx in range(2, min(worksheet.max_row + 1, 101)):  # Sample up to 100 rows
        for column_idx in range(1, worksheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(column_idx)
            cell = worksheet[f"{col_letter}{row_idx}"]
            column_name = column_names.get(column_idx)
            
            if column_name and cell.value is not None:
                # Determine the type of this cell value
                if isinstance(cell.value, int):
                    cell_type = "int"
                elif isinstance(cell.value, float):
                    cell_type = "float"
                elif isinstance(cell.value, str):
                    # Further analyze string to check if it could be a date or other format
                    if cell.value.startswith("$") and cell.value[1:].replace('.', '', 1).isdigit():
                        cell_type = "currency"
                    elif cell.value.endswith("%") and cell.value[:-1].replace('.', '', 1).isdigit():
                        cell_type = "percentage"
                    else:
                        cell_type = "string"
                else:
                    cell_type = "string"
                
                # Update column type based on precedence
                current_type = column_types[column_name]
                if current_type == "string" or (
                    current_type == "int" and cell_type in ["float", "currency", "percentage"]
                ):
                    column_types[column_name] = cell_type
    
    logger.debug(f"Detected data types for {len(column_types)} columns")
    return column_types


def freeze_header_pane(worksheet: openpyxl.worksheet.worksheet.Worksheet, freeze_header: bool) -> None:
    """
    Freezes the first row of the worksheet so headers remain visible during scrolling.
    
    Args:
        worksheet: The openpyxl worksheet to modify
        freeze_header: Boolean flag to determine if header should be frozen
        
    Returns:
        None: Modifies worksheet in-place
    """
    if not freeze_header:
        logger.debug("Header freezing skipped (freeze_header=False)")
        return
    
    # Freeze the first row
    worksheet.freeze_panes = 'A2'
    logger.debug("Applied header freeze at row 1")


class ColumnFormatter:
    """
    A class that handles Excel column formatting during the conversion process, including header styling,
    width adjustment, and data type formatting.
    """
    
    def __init__(self, options: ExcelOptions, json_data: Optional[JSONData] = None):
        """
        Initializes a new ColumnFormatter instance with the provided options and JSON data.
        
        Args:
            options: Excel formatting options
            json_data: Optional JSON data structure for enhanced formatting decisions
        """
        self._options = options
        self._json_data = json_data
        self._logger = get_logger(__name__)
        self._logger.debug("ColumnFormatter initialized")
    
    def format_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Applies all configured formatting options to an Excel worksheet.
        
        Args:
            worksheet: The worksheet to format
            
        Returns:
            None: Modifies worksheet in-place
        """
        self._logger.info("Starting worksheet formatting")
        
        # Apply formatting options in a logical order
        self.format_headers(worksheet)
        self.adjust_widths(worksheet)
        self.apply_formatting(worksheet)
        
        self._logger.info("Worksheet formatting completed")
    
    def format_headers(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Formats the header row of the worksheet based on configuration options.
        
        Args:
            worksheet: The worksheet to format
            
        Returns:
            None: Modifies worksheet in-place
        """
        self._logger.debug("Formatting headers")
        format_column_headers(worksheet, self._options.format_headers)
        freeze_header_pane(worksheet, self._options.freeze_header_row)
    
    def adjust_widths(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Adjusts column widths based on content and configuration options.
        
        Args:
            worksheet: The worksheet to adjust
            
        Returns:
            None: Modifies worksheet in-place
        """
        self._logger.debug("Adjusting column widths")
        adjust_column_widths(worksheet, self._options.auto_column_width)
    
    def apply_formatting(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Applies data type-specific formatting to cells based on content and configuration.
        
        Args:
            worksheet: The worksheet to format
            
        Returns:
            None: Modifies worksheet in-place
        """
        self._logger.debug("Applying data formatting")
        column_formats = self.get_column_formats()
        apply_data_formatting(worksheet, column_formats, self._options.apply_data_formatting)
    
    def get_column_formats(self) -> Dict[str, str]:
        """
        Retrieves column format specifications from options and JSON schema.
        
        Returns:
            A dictionary mapping column names to format strings
        """
        # Start with explicit formats from options
        formats = dict(self._options.column_formats or {})
        
        # Enhance with formats based on JSON schema if available
        if self._json_data:
            schema = self._json_data.get_schema()
            if isinstance(schema, dict) and schema.get('type') == 'object' and 'properties' in schema:
                for prop, details in schema.get('properties', {}).items():
                    prop_type = details.get('type')
                    if prop not in formats and prop_type:
                        # Map JSON schema types to Excel formats
                        if prop_type == 'integer':
                            formats[prop] = DATA_TYPE_FORMATS['int']
                        elif prop_type == 'number':
                            formats[prop] = DATA_TYPE_FORMATS['float']
                        elif prop_type == 'string' and 'format' in details:
                            # Handle special string formats
                            if details['format'] == 'date-time':
                                formats[prop] = DATA_TYPE_FORMATS['datetime']
                            elif details['format'] == 'date':
                                formats[prop] = DATA_TYPE_FORMATS['date']
        
        return formats